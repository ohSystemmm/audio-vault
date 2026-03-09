import os
import openpyxl
import warnings
from src.models import Song

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')


def load_and_validate(filename):
    """Fast loader — reads rows from Excel with no HTTP checks or interactive prompts."""
    if not os.path.exists(filename):
        print(f"File '{filename}' does not exist.")
        return None

    try:
        wb = openpyxl.load_workbook(filename, data_only=True, read_only=True)
    except Exception as e:
        print(f"Failed to open excel file: {e}")
        return None

    sheet_names = wb.sheetnames
    if not sheet_names:
        print("No sheets found.")
        return None

    sheet = wb[sheet_names[0]]
    rows = list(sheet.iter_rows(values_only=True))
    wb.close()

    if len(rows) <= 1:
        print("No data found in excel file.")
        return None

    songs = []
    for row_vals in rows[1:]:  # skip header row
        vals = list(row_vals)
        while len(vals) < 6:
            vals.append(None)

        def clean(v):
            return str(v).strip() if v is not None else ""

        idx_str = clean(vals[0])
        name    = clean(vals[1])
        artist  = clean(vals[2])
        length  = clean(vals[3])
        yt_link = clean(vals[4])
        sp_link = clean(vals[5])

        try:
            parsed_idx = int(idx_str)
        except (ValueError, TypeError):
            parsed_idx = 0

        songs.append(Song(
            index=parsed_idx,
            name=name,
            artist=artist,
            length=length,
            youtube_link=yt_link,
            spotify_link=sp_link,
        ))

    return songs


def create_excel_from_tracks(filename, tracks):
    """
    Generate a new Excel file from a list of track dicts.
    Columns: IndexNumber | SongName | ArtistName | SongLength | YouTubeLink (blank) | SpotifySongLink
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Songs"

    headers = ["IndexNumber", "SongName", "ArtistName", "SongLength", "YouTubeLink", "SpotifySongLink"]
    ws.append(headers)

    for i, track in enumerate(tracks, start=1):
        ws.append([
            i,
            track.get('name', ''),
            track.get('artist', ''),
            track.get('duration', ''),
            '',                           # YouTubeLink — blank
            track.get('spotify_url', ''),
        ])

    try:
        wb.save(filename)
        print(f"Excel file saved: {filename}  ({len(tracks)} tracks)")
    except Exception as e:
        print(f"Failed to save Excel file: {e}")
